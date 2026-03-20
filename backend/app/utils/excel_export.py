"""
Styled Excel export utilities.
"""
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_styled_excel(df: pd.DataFrame, output_path: Path) -> None:
    """
    Export DataFrame to Excel with professional formatting:
    - Green header row with white bold text
    - Auto-adjusted column widths
    - Cell borders
    - Frozen header row
    """
    df.to_excel(output_path, index=False, sheet_name="Data", engine="openpyxl")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Data"]

    header_fill = PatternFill(
        start_color="217346", end_color="217346", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    ws.freeze_panes = "A2"
    wb.save(output_path)
